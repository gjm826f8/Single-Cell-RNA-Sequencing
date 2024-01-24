import os
from openpyxl import load_workbook
import pandas as pd

file_name_array = []
file_sheet_map = {}
df_map = {}
df_array = []

def getFileFromPath(st, folder_path):
    file_name_array.clear()
    # traverse root directory, and list directories as dirs and files as files
    for root, dirs, files in os.walk(folder_path):
        path = root.split(os.sep)
        # st.sidebar.write((len(path) - len(folder_path.split(os.sep))) * '---', os.path.basename(root))
        for file in files:
            # file extension is xlsx and file name not begin with ./~
            if file.endswith(".xlsx") and not file.startswith(".") and not file.startswith("~"):
                # st.sidebar.write((len(path) - len(folder_path.split(os.sep)) + 1) * '---', file)
                file_name_array.append(root + "\\" + file)
    for file_name in file_name_array:
        sheet_able_map = {}
        for sheet in get_sheetnames_xlsx(file_name):
            sheet_able_map[sheet] = False
        file_sheet_map[file_name] = sheet_able_map
    return file_name_array


def get_sheetnames_xlsx(filepath):
    wb = load_workbook(filepath, read_only=True, keep_links=False)
    return wb.sheetnames

def get_file_sheet_map(sheet_select_map):
    sheet_count = 0
    for select_item in sheet_select_map:
        for select_sheet_name in sheet_select_map[select_item]:
            file_sheet_map[select_item][select_sheet_name] = True
            sheet_count += 1
    print("->>>4: ", file_sheet_map)
    return sheet_count


def read_file(file_path):
    idx = 0
    df_array.clear()
    # read all files in the file_name_array
    for file_name in file_name_array:
        for sheet_name in file_sheet_map[file_name]:
            if file_sheet_map[file_name][sheet_name]:
                df = pd.read_excel(file_name, sheet_name=sheet_name)
                df_array.append(df)
                # df_map[idx] = file_name + " " + sheet_name
                df_map[idx] = file_name
                idx += 1
                print(file_name, sheet_name, df.shape)
        # idx = (int(idx / 100) + 1) * 100
    return df_array